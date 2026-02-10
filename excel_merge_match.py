from __future__ import annotations

import os
import sys
import datetime as _dt
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QApplication,
    QFileDialog,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QFormLayout,
    QGroupBox,
    QLabel,
    QLineEdit,
    QPushButton,
    QComboBox,
    QListWidget,
    QListWidgetItem,
    QAbstractItemView,
    QCheckBox,
    QMessageBox,
)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _try_to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def list_sheets(xlsx_path: str) -> List[str]:
    wb = load_workbook(xlsx_path, data_only=True)
    return list(wb.sheetnames)


def _header_names(ws: Worksheet, header_row: int = 1) -> List[str]:
    names: List[str] = []
    for c in range(1, ws.max_column + 1):
        name = _cell_to_str(ws.cell(row=header_row, column=c).value)
        if not name:
            name = f"列{c}"
        names.append(name)
    return names


def list_columns(xlsx_path: str, sheet_name: str, header_row: int = 1) -> List[str]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    return _header_names(ws, header_row=header_row)


def _get_header_map(ws: Worksheet, header_row: int = 1) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = _cell_to_str(ws.cell(row=header_row, column=c).value)
        if not name:
            name = f"列{c}"
        if name not in m:
            m[name] = c
    return m


class _MergedValueResolver:
    def __init__(self, ws: Worksheet):
        self._lookup: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for r in ws.merged_cells.ranges:
            tl = (r.min_row, r.min_col)
            for row in range(r.min_row, r.max_row + 1):
                for col in range(r.min_col, r.max_col + 1):
                    self._lookup[(row, col)] = tl

    def get(self, ws: Worksheet, row: int, col: int) -> Any:
        tl = self._lookup.get((row, col))
        if tl is None:
            return ws.cell(row=row, column=col).value
        return ws.cell(row=tl[0], column=tl[1]).value


def _build_key(ws: Worksheet, row: int, key_col_indexes: List[int], resolver: _MergedValueResolver) -> str:
    parts: List[str] = []
    for c in key_col_indexes:
        v = resolver.get(ws, row, c)
        s = _cell_to_str(v)
        if s == "":
            return ""
        parts.append(s)
    return "_".join(parts)


def _accumulate(prev: Any, cur: Any) -> Any:
    p = _try_to_float(prev)
    c = _try_to_float(cur)
    if p is not None and c is not None:
        s = p + c
        return int(s) if float(s).is_integer() else s
    prev_s = _cell_to_str(prev)
    cur_s = _cell_to_str(cur)
    if prev_s == "":
        return cur_s
    if cur_s == "":
        return prev_s
    seen = set([x.strip() for x in prev_s.split(";") if x.strip() != ""])
    return prev_s if cur_s in seen else (prev_s + ";" + cur_s)


@dataclass(frozen=True)
class SourceConfig:
    xlsx_path: str
    sheet_name: str
    key_columns: Tuple[str, ...]
    value_column: str
    accumulate: bool = False
    header_row: int = 1


@dataclass(frozen=True)
class TargetConfig:
    xlsx_path: str
    sheet_name: str
    key_columns: Tuple[str, ...]
    write_to_column: Optional[str] = None
    header_row: int = 1


def build_source_mapping(cfg: SourceConfig) -> Dict[str, Any]:
    wb = load_workbook(cfg.xlsx_path, data_only=True)
    ws = wb[cfg.sheet_name]
    header_map = _get_header_map(ws, header_row=cfg.header_row)
    missing = [c for c in (*cfg.key_columns, cfg.value_column) if c not in header_map]
    if missing:
        raise ValueError(f"数据源缺少列：{', '.join(missing)}")
    resolver = _MergedValueResolver(ws)
    key_col_indexes = [header_map[c] for c in cfg.key_columns]
    value_col_index = header_map[cfg.value_column]
    mapping: Dict[str, Any] = {}
    for r in range(cfg.header_row + 1, ws.max_row + 1):
        key = _build_key(ws, r, key_col_indexes, resolver)
        if key == "":
            continue
        value = resolver.get(ws, r, value_col_index)
        if value is None or _cell_to_str(value) == "":
            continue
        if not cfg.accumulate:
            if key not in mapping:
                mapping[key] = value
            continue
        prev = mapping.get(key)
        if prev is None:
            mapping[key] = value
            continue
        mapping[key] = _accumulate(prev, value)
    return mapping


def _ensure_dir(p: str) -> None:
    d = os.path.dirname(os.path.abspath(p))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def apply_mapping_to_target(
    src_cfg: SourceConfig,
    mapping: Dict[str, Any],
    tgt_cfg: TargetConfig,
    output_path: Optional[str] = None,
    output_header: Optional[str] = None,
) -> Tuple[int, int, str]:
    wb = load_workbook(tgt_cfg.xlsx_path)
    ws = wb[tgt_cfg.sheet_name]
    header_map = _get_header_map(ws, header_row=tgt_cfg.header_row)
    write_col_index: Optional[int] = None
    if tgt_cfg.write_to_column:
        if tgt_cfg.write_to_column not in header_map:
            raise KeyError(f"匹配写入列不存在：{tgt_cfg.write_to_column}")
        write_col_index = header_map[tgt_cfg.write_to_column]
    else:
        write_col_index = ws.max_column + 1
        hdr = output_header if output_header and output_header.strip() else f"匹配_{src_cfg.value_column}"
        ws.cell(row=tgt_cfg.header_row, column=write_col_index).value = hdr
    resolver = _MergedValueResolver(ws)
    key_col_indexes = [header_map[c] for c in tgt_cfg.key_columns]
    matched = 0
    total = 0
    for r in range(tgt_cfg.header_row + 1, ws.max_row + 1):
        key = _build_key(ws, r, key_col_indexes, resolver)
        if key == "":
            continue
        total += 1
        v = mapping.get(key)
        if v is not None:
            ws.cell(row=r, column=write_col_index).value = v
            matched += 1
    if not output_path or output_path.strip() == "":
        src = Path(tgt_cfg.xlsx_path)
        output_path = str(src.with_name(f"{src.stem}_matched{src.suffix}"))
    _ensure_dir(output_path)
    wb.save(output_path)
    return matched, total, output_path


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("合并单元格匹配工具")
        cw = QWidget()
        self.setCentralWidget(cw)
        root = QVBoxLayout()
        cw.setLayout(root)
        src_group = QGroupBox("数据源（含合并单元格）")
        tgt_group = QGroupBox("待匹配数据源")
        out_group = QGroupBox("输出")
        root.addWidget(src_group)
        root.addWidget(tgt_group)
        root.addWidget(out_group)
        src_layout = QFormLayout()
        tgt_layout = QFormLayout()
        out_layout = QFormLayout()
        src_group.setLayout(src_layout)
        tgt_group.setLayout(tgt_layout)
        out_group.setLayout(out_layout)
        self.src_path = QLineEdit()
        self.src_browse = QPushButton("选择Excel")
        src_path_row = QHBoxLayout()
        src_path_row.addWidget(self.src_path, 1)
        src_path_row.addWidget(self.src_browse)
        src_layout.addRow(QLabel("文件"), src_path_row)
        self.src_sheet = QComboBox()
        src_layout.addRow(QLabel("Sheet"), self.src_sheet)
        self.src_keys = QListWidget()
        self.src_keys.setSelectionMode(QAbstractItemView.MultiSelection)
        src_keys_row = QHBoxLayout()
        src_keys_row.addWidget(self.src_keys)
        self.src_acc = QCheckBox("累加 Value")
        src_keys_btns = QVBoxLayout()
        btn_src_sel_all = QPushButton("全选")
        btn_src_clear = QPushButton("清空")
        src_keys_btns.addWidget(btn_src_sel_all)
        src_keys_btns.addWidget(btn_src_clear)
        src_keys_btns.addStretch(1)
        src_keys_row.addLayout(src_keys_btns)
        src_layout.addRow(QLabel("Key 列"), src_keys_row)
        self.src_value = QComboBox()
        src_value_row = QHBoxLayout()
        src_value_row.addWidget(self.src_value, 1)
        src_value_row.addWidget(self.src_acc)
        src_layout.addRow(QLabel("Value 列"), src_value_row)
        self.tgt_path = QLineEdit()
        self.tgt_browse = QPushButton("选择Excel")
        tgt_path_row = QHBoxLayout()
        tgt_path_row.addWidget(self.tgt_path, 1)
        tgt_path_row.addWidget(self.tgt_browse)
        tgt_layout.addRow(QLabel("文件"), tgt_path_row)
        self.tgt_sheet = QComboBox()
        tgt_layout.addRow(QLabel("Sheet"), self.tgt_sheet)
        self.tgt_keys = QListWidget()
        self.tgt_keys.setSelectionMode(QAbstractItemView.MultiSelection)
        tgt_keys_row = QHBoxLayout()
        tgt_keys_row.addWidget(self.tgt_keys)
        tgt_keys_btns = QVBoxLayout()
        btn_tgt_sel_all = QPushButton("全选")
        btn_tgt_clear = QPushButton("清空")
        tgt_keys_btns.addWidget(btn_tgt_sel_all)
        tgt_keys_btns.addWidget(btn_tgt_clear)
        tgt_keys_btns.addStretch(1)
        tgt_keys_row.addLayout(tgt_keys_btns)
        tgt_layout.addRow(QLabel("Key 列"), tgt_keys_row)
        self.tgt_write_col = QComboBox()
        tgt_layout.addRow(QLabel("匹配写入列"), self.tgt_write_col)
        self.out_path = QLineEdit()
        self.out_browse = QPushButton("选择输出文件")
        out_path_row = QHBoxLayout()
        out_path_row.addWidget(self.out_path, 1)
        out_path_row.addWidget(self.out_browse)
        out_layout.addRow(QLabel("输出文件"), out_path_row)
        self.run_btn = QPushButton("开始匹配")
        root.addWidget(self.run_btn)
        self.src_browse.clicked.connect(self._on_src_browse)
        self.tgt_browse.clicked.connect(self._on_tgt_browse)
        self.out_browse.clicked.connect(self._on_out_browse)
        self.src_sheet.currentIndexChanged.connect(self._refresh_src_columns)
        self.tgt_sheet.currentIndexChanged.connect(self._refresh_tgt_columns)
        btn_src_sel_all.clicked.connect(lambda: self._select_all(self.src_keys, True))
        btn_src_clear.clicked.connect(lambda: self._select_all(self.src_keys, False))
        btn_tgt_sel_all.clicked.connect(lambda: self._select_all(self.tgt_keys, True))
        btn_tgt_clear.clicked.connect(lambda: self._select_all(self.tgt_keys, False))
        self.run_btn.clicked.connect(self._on_run)

    def _select_all(self, lw: QListWidget, sel: bool) -> None:
        for i in range(lw.count()):
            it = lw.item(i)
            it.setSelected(sel)

    def _on_src_browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "选择数据源Excel", "", "Excel 文件 (*.xlsx)")
        if not path:
            return
        self.src_path.setText(path)
        self._populate_sheets(self.src_sheet, path)
        self._refresh_src_columns()

    def _on_tgt_browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "选择待匹配Excel", "", "Excel 文件 (*.xlsx)")
        if not path:
            return
        self.tgt_path.setText(path)
        self._populate_sheets(self.tgt_sheet, path)
        self._refresh_tgt_columns()
        if self.out_path.text().strip() == "":
            p = Path(path)
            self.out_path.setText(str(p.with_name(f"{p.stem}_matched{p.suffix}")))

    def _on_out_browse(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, "选择输出文件", self.out_path.text().strip() or "", "Excel 文件 (*.xlsx)")
        if not path:
            return
        self.out_path.setText(path)

    def _populate_sheets(self, combo: QComboBox, xlsx_path: str) -> None:
        try:
            sheets = list_sheets(xlsx_path)
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))
            return
        combo.blockSignals(True)
        combo.clear()
        for s in sheets:
            combo.addItem(s)
        combo.blockSignals(False)

    def _refresh_src_columns(self) -> None:
        path = self.src_path.text().strip()
        sheet = self.src_sheet.currentText().strip()
        self.src_keys.clear()
        self.src_value.clear()
        if not path or not sheet:
            return
        try:
            cols = list_columns(path, sheet)
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))
            return
        for c in cols:
            it = QListWidgetItem(c)
            self.src_keys.addItem(it)
            self.src_value.addItem(c)

    def _refresh_tgt_columns(self) -> None:
        path = self.tgt_path.text().strip()
        sheet = self.tgt_sheet.currentText().strip()
        self.tgt_keys.clear()
        self.tgt_write_col.clear()
        if not path or not sheet:
            return
        try:
            cols = list_columns(path, sheet)
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))
            return
        for c in cols:
            it = QListWidgetItem(c)
            self.tgt_keys.addItem(it)
        self.tgt_write_col.addItem("新增到最后一列")
        for c in cols:
            self.tgt_write_col.addItem(c)

    def _selected_columns(self, lw: QListWidget) -> Tuple[str, ...]:
        cols: List[str] = []
        for it in lw.selectedItems():
            cols.append(it.text())
        return tuple(cols)

    def _on_run(self) -> None:
        try:
            src_path = self.src_path.text().strip()
            tgt_path = self.tgt_path.text().strip()
            if not src_path or not tgt_path:
                raise ValueError("请先选择数据源与待匹配文件")
            src_sheet = self.src_sheet.currentText().strip()
            tgt_sheet = self.tgt_sheet.currentText().strip()
            if not src_sheet or not tgt_sheet:
                raise ValueError("请选择两个文件的 Sheet")
            src_keys = self._selected_columns(self.src_keys)
            tgt_keys = self._selected_columns(self.tgt_keys)
            if not src_keys:
                raise ValueError("请选择数据源的 Key 列")
            if not tgt_keys:
                raise ValueError("请选择待匹配的 Key 列")
            val_col = self.src_value.currentText().strip()
            acc = self.src_acc.isChecked()
            src_cfg = SourceConfig(
                xlsx_path=src_path,
                sheet_name=src_sheet,
                key_columns=src_keys,
                value_column=val_col,
                accumulate=acc,
            )
            mapping = build_source_mapping(src_cfg)
            write_to = self.tgt_write_col.currentText().strip()
            if write_to == "新增到最后一列":
                write_to = None
            tgt_cfg = TargetConfig(
                xlsx_path=tgt_path,
                sheet_name=tgt_sheet,
                key_columns=tgt_keys,
                write_to_column=write_to,
            )
            out_path = self.out_path.text().strip()
            matched, total, out_file = apply_mapping_to_target(
                src_cfg=src_cfg,
                mapping=mapping,
                tgt_cfg=tgt_cfg,
                output_path=out_path,
                output_header=None,
            )
            QMessageBox.information(self, "完成", f"匹配成功：{matched}/{total}\n输出文件：{out_file}")
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))


def main() -> None:
    app = QApplication(sys.argv)
    w = MainWindow()
    w.resize(900, 700)
    w.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

